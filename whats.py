from flask import Flask, render_template, request
import openpyxl
from selenium import webdriver
import pyautogui

app = Flask(__name__)

@app.route("/", methods=["GET", "POST"])
def index():
    if request.method == "POST":
        file = request.files["file"]
        # Use Openpyxl to read the Excel file and extract the contacts and messages
        wb = openpyxl.load_workbook(file)
        sheet = wb.active
        contacts = []
        messages = []
        for row in range(1, sheet.max_row + 1):
            contacts.append(sheet.cell(row=row, column=1).value)
            messages.append(sheet.cell(row=row, column=2).value)
        return render_template("index.html", contacts=contacts, messages=messages)
    return render_template("index.html")

@app.route("/send", methods=["POST"])
def send_messages():
    # Use Selenium or pywhatkit to automate sending messages to the contacts
    pyautogui.hotkey('win', 'r')
    pyautogui.typewrite("whatsapp")
    pyautogui.hotkey('enter')
    contacts = request.form.getlist("contacts")
    messages = request.form.getlist("messages")
    for i in range(len(contacts)):
        # Search for the contact and open the chat
        pyautogui.typewrite("/")
        pyautogui.typewrite(contacts[i])
        pyautogui.hotkey('enter')
        # Type the message and send it
        pyautogui.typewrite(messages[i])
        pyautogui.hotkey('enter')
        # send message to contacts[i] with message messages[i]"
    return "Messages sent!"


if __name__ == "__main__":
    app.run(debug=True)
